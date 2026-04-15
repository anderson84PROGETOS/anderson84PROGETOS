import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import psutil
import os
import requests
import hashlib
import subprocess
import webbrowser
from openpyxl import Workbook


class NetstatGUI:

    def __init__(self, root):
        self.root = root
        self.root.title("Netstat Monitor PRO")
        self.root.state("zoomed")

        style = ttk.Style()
        style.theme_use("clam")

        # 🔍 BUSCA
        top_frame = tk.Frame(root)
        top_frame.pack(fill="x", padx=10, pady=5)

        tk.Label(top_frame, text="Buscar:").pack(side="left")

        self.busca_entry = tk.Entry(top_frame, width=30)
        self.busca_entry.pack(side="left", padx=5)
        self.busca_entry.bind("<KeyRelease>", lambda e: self.filtrar())

        tk.Label(top_frame, text="Status:").pack(side="left", padx=10)

        self.status_var = tk.StringVar()
        self.status_combo = ttk.Combobox(
            top_frame,
            textvariable=self.status_var,
            width=30,
            values=[
                "TODOS",
                "ESCUTANDO",
                "ESTABELECIDA",
                "TEMPO_ESPERA",
                "AGUARDANDO_FECHAMENTO",
                "CONECTANDO",
                "RECEBENDO_CONEXAO",
                "FINALIZANDO_1",
                "FINALIZANDO_2",
                "ULTIMO_ACK",
                "FECHANDO",
                "SEM_STATUS"
            ]
        )
        self.status_combo.current(0)
        self.status_combo.pack(side="left")
        self.status_combo.bind("<<ComboboxSelected>>", lambda e: self.filtrar())

        # 📊 TABELA
        tabela_frame = tk.Frame(root)
        tabela_frame.pack(fill="both", expand=True)

        # 🔥 SCROLL
        scrollbar_y = tk.Scrollbar(tabela_frame, orient="vertical")
        scrollbar_y.pack(side="right", fill="y")

        scrollbar_x = tk.Scrollbar(tabela_frame, orient="horizontal")
        scrollbar_x.pack(side="bottom", fill="x")

        self.tree = ttk.Treeview(
            tabela_frame,
            columns=("nome", "ip", "porta", "pid", "status", "class", "pais", "caminho"),
            show="headings",
            yscrollcommand=scrollbar_y.set,
            xscrollcommand=scrollbar_x.set
        )

        scrollbar_y.config(command=self.tree.yview)
        scrollbar_x.config(command=self.tree.xview)

        colunas = ["Processo", "IP", "Porta", "PID", "Status", "Classificação", "Localização", "Caminho"]

        for i, col in enumerate(colunas):
            self.tree.heading(self.tree["columns"][i], text=col)

        self.tree.column("nome", width=150)
        self.tree.column("ip", width=120)
        self.tree.column("porta", width=80)
        self.tree.column("pid", width=80)
        self.tree.column("status", width=150)
        self.tree.column("class", width=120)
        self.tree.column("pais", width=200)
        self.tree.column("caminho", width=500)

        self.tree.pack(fill="both", expand=True)

        self.tree.tag_configure("suspeito", background="#ff9999")
        self.tree.tag_configure("normal", background="#ccffcc")

        # 🔥 CLICK = COPIAR
        self.tree.bind("<ButtonRelease-1>", self.copiar_automatico)
        self.tree.bind("<<TreeviewSelect>>", self.mostrar_caminho)

        # 📄 CAMINHO
        self.caminho_label = tk.Label(root, text="Caminho:", anchor="w")
        self.caminho_label.pack(fill="x", padx=10, pady=5)

        # 🔘 BOTÕES
        botoes = tk.Frame(root)
        botoes.pack(pady=5)

        tk.Button(botoes, text="📋 Copiar Caminho", bg="#00d4ff", command=self.copiar_caminho).grid(row=0, column=0, padx=5)
        tk.Button(botoes, text="💾 TXT", bg="#ff9800", command=self.salvar_txt).grid(row=0, column=1, padx=5)
        tk.Button(botoes, text="📊 Excel", bg="#6a5acd", fg="white", command=self.salvar_excel).grid(row=0, column=2, padx=5)
        tk.Button(botoes, text="🔄 Atualizar", bg="#00ff88", command=self.atualizar).grid(row=0, column=3, padx=5)
        tk.Button(botoes, text="📂 Pasta", bg="#ff6666", command=self.abrir_local).grid(row=0, column=4, padx=5)
        tk.Button(botoes, text="🛡️ VirusTotal", bg="#ffcc00", command=self.abrir_virustotal).grid(row=0, column=5, padx=5)

        self.caminho_processo = ""
        self.dados = []

        self.carregar()

    # 🌍 IP INFO
    def get_ip_info(self, ip):
        if not ip or ip == "127.0.0.1":
            return ""
        try:
            r = requests.get(f"http://ip-api.com/json/{ip}", timeout=2).json()
            return f"{r.get('country','')} / {r.get('city','')}"
        except:
            return "Desconhecido"

    # 🔄 STATUS
    def traduzir_status(self, status):
        mapa = {
            "LISTEN": "ESCUTANDO",
            "ESTABLISHED": "ESTABELECIDA",
            "TIME_WAIT": "TEMPO_ESPERA",
            "CLOSE_WAIT": "AGUARDANDO_FECHAMENTO",
            "SYN_SENT": "CONECTANDO",
            "SYN_RECV": "RECEBENDO_CONEXAO",
            "FIN_WAIT1": "FINALIZANDO_1",
            "FIN_WAIT2": "FINALIZANDO_2",
            "LAST_ACK": "ULTIMO_ACK",
            "CLOSING": "FECHANDO",
            "NONE": "SEM_STATUS"
        }
        return mapa.get(status, status)

    # 🔐 HASH
    def gerar_hash(self, caminho):
        try:
            with open(caminho, "rb") as f:
                return hashlib.sha256(f.read()).hexdigest()
        except:
            return None

    # 🛡️ VIRUSTOTAL
    def abrir_virustotal(self):
        if not self.caminho_processo or self.caminho_processo == "N/A":
            return

        hash_sha256 = self.gerar_hash(self.caminho_processo)
        if not hash_sha256:
            return

        url = f"https://www.virustotal.com/gui/file/{hash_sha256}/detection"

        try:
            subprocess.Popen(['chrome', '--incognito', url])
        except:
            webbrowser.open(url)

    # 📡 CARREGAR
    def carregar(self):
        self.dados.clear()

        for conn in psutil.net_connections(kind='inet'):
            try:
                p = psutil.Process(conn.pid)
                nome = p.name()
                caminho = p.exe()
                pid = conn.pid
            except:
                nome, caminho, pid = "N/A", "N/A", "N/A"

            ip = conn.raddr.ip if conn.raddr else ""
            porta = conn.laddr.port if conn.laddr else ""

            status = self.traduzir_status(conn.status)
            pais = self.get_ip_info(ip)

            classificacao = "Suspeito" if porta not in [80, 443] and conn.status == "ESTABLISHED" else "Normal"

            self.dados.append((nome, ip, porta, pid, status, classificacao, pais, caminho))

        self.atualizar_tabela()

    def atualizar_tabela(self):
        self.tree.delete(*self.tree.get_children())
        for d in self.dados:
            tag = "suspeito" if d[5] == "Suspeito" else "normal"
            self.tree.insert("", "end", values=d, tags=(tag,))

    def filtrar(self):
        busca = self.busca_entry.get().lower()
        status_filtro = self.status_var.get()

        self.tree.delete(*self.tree.get_children())

        for d in self.dados:
            if busca in str(d).lower():
                if status_filtro == "TODOS" or d[4] == status_filtro:
                    tag = "suspeito" if d[5] == "Suspeito" else "normal"
                    self.tree.insert("", "end", values=d, tags=(tag,))

    def atualizar(self):
        self.carregar()

    # 🔥 COPIAR AUTOMÁTICO
    def copiar_automatico(self, event):
        item = self.tree.identify_row(event.y)
        coluna = self.tree.identify_column(event.x)

        if not item:
            return

        valores = self.tree.item(item, "values")
        col_index = int(coluna.replace("#", "")) - 1

        if col_index < len(valores):
            valor = str(valores[col_index])
            self.root.clipboard_clear()
            self.root.clipboard_append(valor)
            messagebox.showinfo("Copiado", "Copiado com sucesso!")

    def mostrar_caminho(self, event):
        item = self.tree.selection()
        if item:
            self.caminho_processo = self.tree.item(item)["values"][7]
            self.caminho_label.config(text=f"Caminho: {self.caminho_processo}")

    def copiar_caminho(self):
        if self.caminho_processo:
            self.root.clipboard_append(self.caminho_processo)

    def abrir_local(self):
        if self.caminho_processo and self.caminho_processo != "N/A":
            os.system(f'explorer /select,"{self.caminho_processo}"')

    def salvar_txt(self):
        caminho = filedialog.asksaveasfilename(defaultextension=".txt")
        if not caminho:
            return

        headers = ["Processo", "IP", "Porta", "PID", "Status", "Classificação", "Localização", "Caminho"]

        # 👇 largura de cada coluna (ajuste se quiser)
        larguras = [20, 40, 20, 20, 30, 20, 38, 40]

        with open(caminho, "w", encoding="utf-8") as f:

            # 🔥 Cabeçalho alinhado
            linha_header = ""
            for i, h in enumerate(headers):
                linha_header += f"{h:<{larguras[i]}}"
            f.write(linha_header + "\n")

            f.write("-" * sum(larguras) + "\n")

            # 🔥 Dados alinhados
            for item in self.tree.get_children():
                valores = self.tree.item(item)["values"]

                linha = ""
                for i, val in enumerate(valores):
                    linha += f"{str(val):<{larguras[i]}}"

                f.write(linha + "\n")

        messagebox.showinfo("Sucesso", "Arquivo TXT salvo estilo Excel!")

    def salvar_excel(self):
        caminho = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if not caminho:
            return

        wb = Workbook()
        ws = wb.active

        headers = ["Processo", "IP", "Porta", "PID", "Status", "Classificação", "Localização", "Caminho"]

        # 🔥 CABEÇALHO
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # 🔥 DADOS
        for row_idx, item in enumerate(self.tree.get_children(), start=2):
            valores = self.tree.item(item)["values"]

            for col_idx, val in enumerate(valores, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        wb.save(caminho)

        messagebox.showinfo("Sucesso", "Arquivo Excel salvo com sucesso!")


if __name__ == "__main__":
    root = tk.Tk()
    app = NetstatGUI(root)
    root.mainloop()
