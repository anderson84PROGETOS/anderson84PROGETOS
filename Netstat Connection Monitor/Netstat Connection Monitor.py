import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import psutil
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class NetstatGUI:

    def __init__(self, root):
        self.root = root
        self.root.title("Netstat Connection Monitor")
        self.root.state("zoomed")

        style = ttk.Style()
        style.theme_use("clam")

        # 🔍 BUSCA
        top_frame = tk.Frame(root)
        top_frame.pack(fill="x", padx=10, pady=5)

        tk.Label(top_frame, text="Buscar:").pack(side="left")

        self.busca_entry = tk.Entry(top_frame, width=30)
        self.busca_entry.pack(side="left", padx=5)
        self.busca_entry.bind("<KeyRelease>", lambda e: self.filtrar())

        tk.Label(top_frame, text="Status:").pack(side="left", padx=10)

        self.status_var = tk.StringVar()
        self.status_combo = ttk.Combobox(
            top_frame,
            textvariable=self.status_var,
            width=30,  # 👈 aumenta aqui
            values=[
                "TODOS",
                "ESCUTANDO",
                "ESTABELECIDA",
                "TEMPO_ESPERA",
                "AGUARDANDO_FECHAMENTO",
                "CONECTANDO",
                "RECEBENDO_CONEXAO",
                "FINALIZANDO_1",
                "FINALIZANDO_2",
                "ULTIMO_ACK",
                "FECHANDO",
                "SEM_STATUS"
            ]
        )
        self.status_combo.current(0)
        self.status_combo.pack(side="left")
        self.status_combo.bind("<<ComboboxSelected>>", lambda e: self.filtrar())

        # 📊 TABELA
        tabela_frame = tk.Frame(root)
        tabela_frame.pack(fill="both", expand=True)

        scrollbar_y = tk.Scrollbar(tabela_frame, orient="vertical")
        scrollbar_y.pack(side="right", fill="y")

        scrollbar_x = tk.Scrollbar(tabela_frame, orient="horizontal")
        scrollbar_x.pack(side="bottom", fill="x")

        self.tree = ttk.Treeview(
            tabela_frame,
            columns=("nome", "porta", "endereco", "pid", "status", "classificacao", "caminho"),
            show="headings",
            yscrollcommand=scrollbar_y.set,
            xscrollcommand=scrollbar_x.set
        )

        scrollbar_y.config(command=self.tree.yview)
        scrollbar_x.config(command=self.tree.xview)

        colunas = ["Processo", "Porta", "Endereço", "PID", "Status", "Classificação", "Caminho"]

        for i, col in enumerate(colunas):
            self.tree.heading(self.tree["columns"][i], text=col)

        self.tree.column("nome", width=180)
        self.tree.column("porta", width=70)
        self.tree.column("endereco", width=200)
        self.tree.column("pid", width=80)
        self.tree.column("status", width=150)
        self.tree.column("classificacao", width=150)
        self.tree.column("caminho", width=800)

        self.tree.pack(fill="both", expand=True)

        # 🎨 CORES
        self.tree.tag_configure("suspeito", background="#ffb3b3")
        self.tree.tag_configure("normal", background="#b3ffb3")

        self.tree.bind("<<TreeviewSelect>>", self.mostrar_caminho)

        # 🆕 DUPLO CLIQUE PARA COPIAR
        self.tree.bind("<Double-1>", self.copiar_celula)

        # 📄 CAMINHO
        self.caminho_label = tk.Label(root, text="Caminho do arquivo:", anchor="w")
        self.caminho_label.pack(fill="x", padx=10, pady=5)

        # 🔘 BOTÕES
        botoes = tk.Frame(root)
        botoes.pack(pady=5)

        tk.Button(botoes, text="Copiar o Caminho", bg="#03e8fc", command=self.copiar_caminho).grid(row=0, column=0, padx=5)
        tk.Button(botoes, text="Salvar TXT", bg="#fc9d03", command=self.salvar_txt).grid(row=0, column=1, padx=5)
        tk.Button(botoes, text="Salvar Excel", bg="#6a5acd", fg="white", command=self.salvar_excel).grid(row=0, column=2, padx=5)
        tk.Button(botoes, text="Atualizar", bg="#03fc0b", command=self.atualizar).grid(row=0, column=3, padx=5)
        tk.Button(botoes, text="Abrir Pasta", bg="#f79494", command=self.abrir_local).grid(row=0, column=4, padx=5)

        self.caminho_processo = ""
        self.dados = []

        self.carregar_conexoes()

    # 🆕 COPIAR CELULA
    def copiar_celula(self, event):
        item = self.tree.identify_row(event.y)
        coluna = self.tree.identify_column(event.x)

        if not item:
            return

        valores = self.tree.item(item, "values")
        col_index = int(coluna.replace("#", "")) - 1

        if col_index < len(valores):
            valor = str(valores[col_index])
            self.root.clipboard_clear()
            self.root.clipboard_append(valor)
            messagebox.showinfo("Copiado", f"Valor copiado:\n{valor}")

    def traduzir_status(self, status):
        mapa = {
            "LISTEN": "ESCUTANDO",
            "ESTABLISHED": "ESTABELECIDA",
            "TIME_WAIT": "TEMPO_ESPERA",
            "CLOSE_WAIT": "AGUARDANDO_FECHAMENTO",
            "SYN_SENT": "CONECTANDO",
            "SYN_RECV": "RECEBENDO_CONEXAO",
            "FIN_WAIT1": "FINALIZANDO_1",
            "FIN_WAIT2": "FINALIZANDO_2",
            "LAST_ACK": "ULTIMO_ACK",
            "CLOSING": "FECHANDO",
            "NONE": "SEM_STATUS"
        }
        return mapa.get(status, status)

    def classificar(self, porta, status_original):
        portas_seguras = [80, 443, 53]

        if porta not in portas_seguras and status_original == "ESTABLISHED":
            return "Suspeito"

        return "Normal"

    def carregar_conexoes(self):
        self.dados.clear()

        for conn in psutil.net_connections(kind='inet'):

            try:
                processo = psutil.Process(conn.pid)
                nome = processo.name()
                caminho = processo.exe()
                pid = conn.pid
            except:
                nome = "N/A"
                caminho = "N/A"
                pid = "N/A"

            porta = conn.laddr.port if conn.laddr else "N/A"
            endereco = conn.laddr.ip if conn.laddr else "N/A"

            status_original = conn.status
            status_traduzido = self.traduzir_status(status_original)

            classificacao = self.classificar(porta, status_original)

            item = (nome, porta, endereco, pid, status_traduzido, classificacao, caminho)
            self.dados.append(item)

        self.atualizar_tabela()

    def atualizar_tabela(self):
        self.tree.delete(*self.tree.get_children())

        for d in self.dados:
            tag = "suspeito" if d[5] == "Suspeito" else "normal"
            self.tree.insert("", "end", values=d, tags=(tag,))

    def filtrar(self):
        busca = self.busca_entry.get().lower()
        status_filtro = self.status_var.get()

        self.tree.delete(*self.tree.get_children())

        for d in self.dados:
            texto = f"{d[0]} {d[1]} {d[3]}".lower()

            if busca in texto:
                if status_filtro == "TODOS" or d[4] == status_filtro:
                    tag = "suspeito" if d[5] == "Suspeito" else "normal"
                    self.tree.insert("", "end", values=d, tags=(tag,))

    def atualizar(self):
        self.carregar_conexoes()

    def mostrar_caminho(self, event):
        item = self.tree.selection()
        if not item:
            return

        dados = self.tree.item(item)["values"]
        self.caminho_processo = dados[6]

        self.caminho_label.config(text=f"Caminho do arquivo: {self.caminho_processo}")

    def copiar_caminho(self):
        if not self.caminho_processo:
            return

        self.root.clipboard_clear()
        self.root.clipboard_append(self.caminho_processo)
        messagebox.showinfo("Sucesso", "Caminho copiado!")

    def abrir_local(self):
        if not self.caminho_processo or self.caminho_processo == "N/A":
            messagebox.showwarning("Aviso", "Nenhum caminho válido!")
            return

        os.system(f'explorer /select,"{self.caminho_processo}"')

    def salvar_txt(self):
        caminho = filedialog.asksaveasfilename(defaultextension=".txt")

        if not caminho:
            return

        with open(caminho, "w", encoding="utf-8") as f:
            for item in self.tree.get_children():
                f.write(str(self.tree.item(item)["values"]) + "\n")

        messagebox.showinfo("Sucesso", "Arquivo salvo!")

    def salvar_excel(self):
        caminho = filedialog.asksaveasfilename(defaultextension=".xlsx")

        if not caminho:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Conexões"

        headers = ["Processo", "Porta", "Endereço", "PID", "Status", "Classificação", "Caminho"]

        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, item in enumerate(self.tree.get_children(), start=2):
            valores = self.tree.item(item)["values"]

            for col_idx, valor in enumerate(valores, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)

                if valores[5] == "Suspeito":
                    cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(caminho)

        messagebox.showinfo("Sucesso", "Arquivo Excel salvo com sucesso!")


if __name__ == "__main__":
    root = tk.Tk()
    app = NetstatGUI(root)
    root.mainloop()
